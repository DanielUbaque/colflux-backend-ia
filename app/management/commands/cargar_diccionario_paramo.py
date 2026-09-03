"""Carga el diccionario de terminos de campo del paramo.

Acepta el Excel original (.xlsx) o el JSON equivalente (.json). Crea un
DocumentoConocimiento con el texto completo (para trazabilidad) y una fila de
TerminoCampo por cada observacion, con su vector de embedding.
"""

import json
import os

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from app.ia.cliente import generar_embeddings
from app.models import DocumentoConocimiento, TerminoCampo

MAPEO_COLUMNAS = {
    "Punto": "categoria",
    "Observacion de campo": "observacion_campo",
    "Observación de campo": "observacion_campo",
    "Definicion ecologica": "definicion_ecologica",
    "Definición ecológica": "definicion_ecologica",
    "Variable asociada": "variable_asociada",
    "Variable secundaria": "variable_secundaria",
    "Regla cuantitativa": "regla_cuantitativa",
    "Interpretacion": "interpretacion",
    "Interpretación": "interpretacion",
}

LIMITES = {
    "categoria": 60,
    "observacion_campo": 255,
    "variable_asociada": 120,
    "variable_secundaria": 255,
}


def _leer_json(ruta):
    with open(ruta, "r", encoding="utf-8") as f:
        datos = json.load(f)
    return datos.get("diccionario", []), datos.get("semaforo", [])


def _leer_xlsx(ruta):
    from openpyxl import load_workbook

    wb = load_workbook(ruta, data_only=True)
    hoja = wb[wb.sheetnames[0]]
    filas = list(hoja.iter_rows(values_only=True))
    if not filas:
        return [], []
    encabezado = [str(c).strip() if c else "" for c in filas[0]]
    registros = []
    for fila in filas[1:]:
        if not any(fila):
            continue
        registros.append({
            k: (str(v).strip() if v is not None else "")
            for k, v in zip(encabezado, fila)
        })
    semaforo = []
    if len(wb.sheetnames) > 1:
        hoja2 = wb[wb.sheetnames[1]]
        filas2 = list(hoja2.iter_rows(values_only=True))
        if filas2:
            enc2 = [str(c).strip() if c else "" for c in filas2[0]]
            for fila in filas2[1:]:
                if not any(fila):
                    continue
                semaforo.append({
                    k: (str(v).strip() if v is not None else "")
                    for k, v in zip(enc2, fila)
                })
    return registros, semaforo


class Command(BaseCommand):
    help = "Carga el diccionario de terminos de campo del paramo con sus embeddings."

    def add_arguments(self, parser):
        parser.add_argument("--archivo", required=True, help="Ruta al .xlsx o .json")
        parser.add_argument(
            "--reemplazar", action="store_true",
            help="Borra los terminos previos de origen 'excel' antes de cargar.",
        )
        parser.add_argument(
            "--sin-embeddings", action="store_true",
            help="Carga los textos sin llamar a la API (util para probar).",
        )

    def handle(self, *args, **opciones):
        ruta = opciones["archivo"]
        if not os.path.exists(ruta):
            raise CommandError(f"No existe el archivo: {ruta}")

        if ruta.lower().endswith(".json"):
            registros, semaforo = _leer_json(ruta)
        elif ruta.lower().endswith((".xlsx", ".xlsm")):
            registros, semaforo = _leer_xlsx(ruta)
        else:
            raise CommandError("El archivo debe ser .json o .xlsx")

        if not registros:
            raise CommandError("El archivo no tiene filas de diccionario.")

        self.stdout.write(f"Filas leidas: {len(registros)} (semaforo: {len(semaforo)})")

        lineas = []
        for r in registros:
            lineas.append(" | ".join(f"{k}: {v}" for k, v in r.items() if v))
        texto_completo = "\n".join(lineas)
        if semaforo:
            texto_completo += "\n\nSEMAFORO CUANTITATIVO\n"
            for s in semaforo:
                texto_completo += " | ".join(f"{k}: {v}" for k, v in s.items() if v) + "\n"

        with transaction.atomic():
            documento, creado = DocumentoConocimiento.objects.update_or_create(
                titulo="Diccionario de terminos de campo - paramo",
                defaults={
                    "tipo": "diccionario",
                    "texto_completo": texto_completo,
                    "archivo_origen": os.path.basename(ruta),
                    "metadatos": {"semaforo": semaforo},
                    "procesado": True,
                },
            )
            self.stdout.write(
                f"Documento {'creado' if creado else 'actualizado'}: id={documento.id}"
            )

            if opciones["reemplazar"]:
                borrados, _ = TerminoCampo.objects.filter(origen="excel").delete()
                self.stdout.write(f"Terminos previos borrados: {borrados}")

            objetos = []
            for r in registros:
                campos = {}
                for clave, valor in r.items():
                    destino = MAPEO_COLUMNAS.get(str(clave).strip())
                    if destino:
                        campos[destino] = valor
                if not campos.get("observacion_campo"):
                    continue
                for campo, limite in LIMITES.items():
                    if campo in campos and len(campos[campo]) > limite:
                        campos[campo] = campos[campo][:limite]
                objetos.append(TerminoCampo(
                    origen="excel",
                    documento_origen=documento,
                    confirmado=True,
                    **campos,
                ))

            TerminoCampo.objects.bulk_create(objetos)
            self.stdout.write(f"Terminos creados: {len(objetos)}")

        if opciones["sin_embeddings"]:
            self.stdout.write(self.style.WARNING("Sin embeddings (--sin-embeddings)."))
            return

        pendientes = list(
            TerminoCampo.objects.filter(documento_origen=documento, embedding__isnull=True)
        )
        self.stdout.write(f"Generando embeddings para {len(pendientes)} terminos...")
        textos = [t.texto_para_embedding() for t in pendientes]
        vectores = generar_embeddings(textos)
        for termino, vector in zip(pendientes, vectores):
            termino.embedding = vector
        TerminoCampo.objects.bulk_update(pendientes, ["embedding"], batch_size=50)

        self.stdout.write(self.style.SUCCESS(
            f"Listo. {len(pendientes)} terminos con embedding de 768 dimensiones."
        ))
